import glob
import json
import os
import pandas as pd
import re
import shutil
import uuid

from collections import namedtuple
from datetime import datetime
from openpyxl.utils import get_column_letter
from pathlib import Path

# Helpers: Variables
DWC_NAMESPACE_PREFIXES = ["dwc", "dcterms"]
DEFAULT_SCHEMA_EXTENSION = ".xlsx"

SCHEMA_DIR_PATH = "schemas"

SCHEMA_FILE_PATH = next(
    (
        str(file)
        for file in Path(SCHEMA_DIR_PATH).iterdir()
        if file.is_file()
        and file.name.endswith(DEFAULT_SCHEMA_EXTENSION)
        and file.name.startswith("singlecell")
    ),
    None,
)
# Extract filename without extension
SCHEMA_FILENAME_WO_EXT = Path(
    SCHEMA_FILE_PATH
).stem  # e.g. 'singlecell_schema_main_v0.1'
# Find the part starting with 'v' at the end
SCHEMA_VERSION = (
    SCHEMA_FILENAME_WO_EXT.split("_")[-1]
    if SCHEMA_FILENAME_WO_EXT.split("_")[-1].startswith("v")
    else None
)

# Global variables
CHECKLISTS_DICT = dict()  # Global set to store the schema names
COMPONENTS = dict()  # Global set to store the components
REGEX_TO_SPREADSHEET_DATA_VALIDATION_MAPPING = dict()

# Helpers: Mappings
FORMATS = {
    "xlsx": ".xlsx",
    "json": ".json",
    "xml": ".xml",
    "html": ".html",
}

# Gives names to each element in a tuple
VersionData = namedtuple(
    "VersionData",
    [
        "technology_name",
        "technology_label",
        "standard_name",
        "standard_label",
        "version_description",
    ],
)


# Helpers: Functions
def apply_data_validation(
    component_df, dataframe, pandas_writer, namespace, allowed_values_dict
):
    column_names = (
        component_df["term_name"].drop_duplicates().tolist()
    )  # Use term name instead of term label

    sheet_name = get_worksheet_info(component_df, return_label=False)
    sheet = pandas_writer.sheets[sheet_name]
    workbook = pandas_writer.book

    # Create a hidden sheet for long dropdown lists
    hidden_sheet_name = "HiddenDropdowns"
    hidden_sheet = workbook.get_worksheet_by_name(hidden_sheet_name)

    if not hidden_sheet:
        hidden_sheet = workbook.add_worksheet(hidden_sheet_name)
        hidden_sheet.hide()  # Hide the worksheet

    # Remove duplicate columns from the DataFrame
    dataframe = dataframe.loc[:, ~dataframe.columns.duplicated()]

    for column_name in column_names:
        term_name = component_df.loc[
            component_df["term_name"] == column_name, "term_name"
        ].iloc[0]
        dropdown_list = allowed_values_dict.get(term_name, [])
        regex = (
            component_df.loc[
                component_df["term_name"] == column_name, "term_regex"
            ].iloc[0]
            if "term_regex" in component_df
            else ""
        )
        term_error_message = (
            component_df.loc[
                component_df["term_name"] == column_name, "term_error_message"
            ].iloc[0]
            if "term_error_message" in component_df
            else ""
        )

        # Get spreadsheet official column header letter
        # Indexing starts at 0 by default but in this case, it should start at 1 so increment by 1
        column_index = column_names.index(column_name)
        column_letter = get_column_letter(column_index + 1)

        # Get first row to the last row in a column
        # NB: The first 4 rows of the sheet are locked so the data starts from row 5
        # row_start = 5 # Start from row 5
        # row_end = 1005 # End at row 1005
        # Start from row 5, End at row 1005
        row_start, row_end = 5, max(1005, len(dataframe) + 5)
        row_start_end = f"{column_letter}{row_start}:{column_letter}{row_end}"

        # Apply data formula to the column if regex is provided
        # Ensure that data that has allowed values/dropdown list is not validated by regex
        if regex and not dropdown_list:
            validation_formula = get_xlsx_data_validation_from_regex(
                regex, column_letter
            )
            if validation_formula:
                validation_dict = {}
                # validation_dict = {"validate": "custom", "value": validation_formula}
                if term_error_message:
                    validation_dict["error_message"] = term_error_message
                sheet.data_validation(row_start_end, validation_dict)

        # Apply the dropdown list to the column
        if dropdown_list:
            dropdown_list = list(set(dropdown_list))  # Remove duplicates

            # Capitalise the first letter of each word in the list and replace underscores with spaces
            dropdown_list = [i.title().replace("_", " ") for i in dropdown_list]
            dropdown_list.sort()  # Sort the list in ascending order
            number_of_characters = len(
                ",".join(dropdown_list)
            )  # Calculate the total length of the string

            if number_of_characters >= 255:
                print(
                    f'Info: "{column_name}" column dropdown too long for spreadsheet. A hidden sheet will be created.'
                )

                # Start from row 5, leave row 1 for header, row 2 for file description, and row 3 for example data
                for index, val in enumerate(dropdown_list, start=row_start):
                    hidden_sheet.write(f"{column_letter}{index}", val)

                # Create a range reference for the hidden sheet
                data_validation_range = f"={hidden_sheet_name}!${column_letter}${row_start}:${column_letter}${index}"
                sheet.data_validation(
                    row_start_end,
                    {
                        "validate": "list",
                        "source": data_validation_range,
                        "input_message": "Choose from the list",
                    },
                )
            else:
                sheet.data_validation(
                    row_start_end,
                    {
                        "validate": "list",
                        "source": dropdown_list,
                        "input_message": "Choose from the list",
                    },
                )


def autofit_all_sheets(writer):
    for sheet in writer.sheets.values():
        sheet.autofit()


def remove_existing_schema_files():
    # Make sure the path ends with a separator
    target_dir = os.path.abspath(SCHEMA_DIR_PATH)

    # Match any schema-like files ending in .json or .yaml/.yml
    json_files = glob.glob(os.path.join(target_dir, "*.json"))
    yaml_files = glob.glob(os.path.join(target_dir, "*.yaml")) + glob.glob(
        os.path.join(target_dir, "*.yml")
    )

    # Define what qualifies as a schema file (adjust as needed)
    matching_json = [
        f for f in json_files if "schema" in f.lower() and "singlecell" in f.lower()
    ]
    matching_yaml = [
        f for f in yaml_files if "schema" in f.lower() and "singlecell" in f.lower()
    ]

    for f in matching_json + matching_yaml:
        try:
            os.remove(f)
            print(f'Deleted existing file: {f}')
        except Exception as e:
            print(f'Could not delete {f}: {e}')


def is_camel_case(text):
    # Regular expression to check if text follows camelCase
    return bool(re.match(r"^[a-z]+(?:[A-Z][a-z]+)*$", text))


def is_title_case_with_spaces(text):
    # Regular expression to check if text follows Title Case
    return bool(re.match(r"^[A-Z][a-z]+(?: [A-Z][a-z]+)*$", text))


def convert_datetime(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()  # Converts datetime to 'YYYY-MM-DDTHH:MM:SS'
    return obj


def convert_string_to_title_case(text):
    """
    Convert a given string to title case, handling camel case by adding spaces
    where necessary and replacing certain abbreviations and terms.
    """
    # Convert camelCase to space-separated words if applicable
    if is_camel_case(text):
        text = re.sub(r"([A-Z])", r" \1", text).strip()

    # Ensure title case format with spaces if not already properly formatted
    if not is_title_case_with_spaces(text):
        text = re.sub(r"(?<!^)(?=[A-Z])", " ", text)

    # Apply title casing and replace certain terms
    return (
        text.title()
        .replace("_", " ")
        .replace("  ", " ")
        .replace("I D", "ID")
        .replace("Geogr", "Geographic")
        .replace("Locat", "Location")
        .replace("Latit", "Latitude")
        .replace("Longi", "Longitude")
        .replace("Longitudegitude", "Longitude")
        .replace("Latitudeitude", "Latitude")
        .replace("Locationation", "Location")
        .replace("Geographicreference", "Geographical Reference")
        .replace("Cdna", "cDNA")
    )


def create_readme_worksheet(readme_sheet_data):
    # Create README sheet
    technology_name = readme_sheet_data["technology_name"]
    technology_label = readme_sheet_data["technology_label"]
    version_description = readme_sheet_data["version_description"]
    standard_name = readme_sheet_data["standard_name"]
    standard_label = readme_sheet_data["standard_label"]
    version_column_name = readme_sheet_data["version_column_name"]
    writer = readme_sheet_data["writer"]
    workbook = writer.book
    locked_format = readme_sheet_data["locked_format"]

    readme_df = pd.DataFrame(
        {
            "key": [version_column_name],
            "name": [f"{technology_label} [{standard_label}]"],
            "description": [version_description],
            "standard": [standard_name],
            "technology": [technology_name],
            "manifest_version": [SCHEMA_VERSION.replace("v", "")],
        }
    )

    # Write README sheet with formatting
    readme_df.to_excel(writer, sheet_name="README", index=False)

    # Get the README worksheet
    readme_worksheet = writer.sheets["README"]

    # Apply formatting to the header row (bold text)
    header_format = workbook.add_format({"bold": True, "locked": True})
    readme_worksheet.set_row(0, None, header_format)

    # Lock all columns dynamically
    num_columns = len(readme_df.columns)
    column_range = f"A:{chr(65 + num_columns - 1)}"  # Calculate the column range dynamically (e.g., A:E)

    # Lock the cells in the range determined by the DataFrame columns
    readme_worksheet.set_column(column_range, None, locked_format)

    # Protect the worksheet to prevent editing
    readme_worksheet.protect(
        password=str(uuid.uuid4())
    )  # This will use a random UUID as the password


def datetime_converter(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()  # Convert datetime to ISO format string
    raise TypeError(f"Type {type(obj)} not serialisable")


def filter_data_frame(element):
    """
    Filters the DataFrame rows based on column headers that matches
    the version column name and non-empty values within that column

    Parameters:
    element (dict): A dictionary containing the following:
        data_df (pd.DataFrame): The input data frame.
        standard_name (str): The namespace prefix to filter by.
        technology_name (str): The schema name to filter by.
        technology_label (str): The schema label to set in the config.
        version_column_name: Name of the column pertaining to the namespace

    Returns:
       - pd.DataFrame: The original DataFrame filtered by rows matching the criteria.
    """
    data_df = element["data_df"]
    standard_name = element["standard_name"]
    technology_name = element["technology_name"]
    technology_label = element["technology_label"]
    version_column_name = element["version_column_name"]

    standard_name = "dwc" if standard_name in DWC_NAMESPACE_PREFIXES else standard_name

    # Filter the DataFrame by the version column name
    if version_column_name not in data_df.columns:
        print(
            f'\nNo data found for {standard_name} in {technology_name} technology and "{version_column_name}" version column.'
        )
        return pd.DataFrame()

    # Filter rows where the version column has non-empty (not null, not empty) values           non_empty = data_df[version_column_name].notna() & (data_df[version_column_name] != '')
    non_empty = data_df[version_column_name].notna() & (
        data_df[version_column_name] != ""
    )

    # Check that the non-empty rows are actually 'M' (mandatory) or 'O' (optional) in the version column
    validate_version_column_value(data_df, non_empty, version_column_name)

    # Return the original DataFrame with only the matching rows
    return data_df[non_empty]


def format_and_protect_worksheet(element):
    """
    This function applies formatting and protection to the given worksheet.

    Parameters:
    element (dict): A dictionary containing the following:
        worksheet: The worksheet to format.
        column_names: List of column names for determining the last column.
        locked_format: The format to lock the cells.
        merge_format: The format for merged cells.
    """
    worksheet = element["worksheet"]
    column_names = element["column_names"]
    required_columns = element["required_columns"]
    col_desc_eg = element["col_desc_eg"]
    locked_format = element["locked_format"]
    unlocked_format = element["unlocked_format"]
    merge_format = element["merge_format"]
    required_format = element["required_format"]
    desc_eg_format = element["desc_eg_format"]

    # Get the lexicographical letter of the last column based on the index
    last_column_letter = get_column_letter(len(column_names))

    # Write header in row 1 (header) and apply formatting
    for col, column_name in enumerate(column_names):
        if column_name in required_columns:
            worksheet.write(
                0, col, column_name, required_format
            )  # Bold the required headers
        else:
            worksheet.write(
                0, col, f"{column_name} (optional)", locked_format
            )  # Add (optional) to non-required headers

    # Write column description on row 2 and example in row 3
    for col, column_name in enumerate(column_names):
        # Write description in row 2 (index 1 in 0-based index)
        # Row 2
        worksheet.write(1, col, col_desc_eg[column_name]["description"], desc_eg_format)

        # Write example in row 3 (index 2 in 0-based index)
        # Row 3
        worksheet.write(
            2, col, f'e.g. {col_desc_eg[column_name]["example"]}', desc_eg_format
        )

    # Merge and write instruction in row 4
    merge_row(worksheet, 4, last_column_letter, merge_format)

    # Set the conditional format for locking rows 1 to 4
    worksheet.conditional_format(
        f"A1:{last_column_letter}4", {"type": "no_errors", "format": locked_format}
    )

    # Set all rows below row 4 to unlocked
    worksheet.set_column(f"A5:{last_column_letter}1005", None, unlocked_format)

    # Protect the worksheet. This will use a random UUID as the password
    worksheet.protect(password=str(uuid.uuid4()))


def get_base_schema_json(element):
    """
    Load data from an spreadsheet file and return JSON data filtered by namespace prefix.
    The base namespace prefix, 'ei', data (i.e., rows that do not match the filters) should be returned in
    addition to the provided inputs.

    Parameters:
        data_df (DataFrame): The DataFrame containing the data from the spreadsheet file.
        allowed_values_dict (dict): A dictionary containing allowed values for each column.
        standard_name (str): The namespace prefix to filter by (optional).

    Returns:
        list: A list of dictionaries representing the filtered JSON data.
    """
    # Generate JSON structure
    json_data = []
    fields_to_check = [
        "copo_name",
        "identifier",
        "referenced_component",
        "term_regex",
        "term_cardinality",
        "term_type",
        "term_reference",
    ]

    data_df = element["data_df"]
    allowed_values_dict = element["allowed_values_dict"]
    technology_name = element["technology_name"]
    technology_label = element["technology_label"]

    for _, row in data_df.iterrows():
        field = {}

        # Add fields dynamically from the dataframe, excluding keys that start with "version_"
        for column in data_df.columns:
            value = row[column]

            # Explicitly convert boolean values to boolean type to avoid JSON serialization issues with '0.0' and '1.0'
            if isinstance(value, bool):
                value = bool(value)
            elif isinstance(value, (int, float)) and value in [1, 0]:
                value = bool(value)

            # Only add to the 'field' dictionary if the value exists i.e. is not empty
            if value:
                field[column] = value
            else:
                # Remove the key if it exists in the dictionary
                field.pop(column, None)

        # Add other fields as strings
        field["technology_name"] = technology_name
        field["technology_label"] = technology_label

        # Conditionally add allowed_values if available and not empty
        allowed_values = allowed_values_dict.get(row["term_name"], [])

        # Validate allowed values
        validate_allowed_values(row, allowed_values)

        if allowed_values and row["term_type"] == "enum":
            allowed_values.sort()  # Sort the allowed values
            field["allowed_values"] = allowed_values

        json_data.append(field)

    return json_data


def get_col_desc_eg(component_df, version_column_name):
    filtered_df = component_df[component_df[version_column_name].isin(["M", "O"])]
    # Use the term name as the key
    return {
        row["term_name"]: {
            "description": row.get("term_description", ""),
            "example": row.get("term_example", ""),
        }
        for _, row in filtered_df.iterrows()
    }


def get_worksheet_info(component_df, return_label=False):
    component_name = component_df["component_name"].iloc[0]

    if component_name not in COMPONENTS:
        raise ValueError(
            f"Component name, '{component_name}', not found in the 'components' worksheet."
        )

    return COMPONENTS[component_name] if return_label else component_name


def get_xlsx_data_validation_from_regex(regex, column_letter):
    # Define a mapping from regex patterns to spreadsheet custom validation formulas
    # NB: Data starts from row 5
    global REGEX_TO_SPREADSHEET_DATA_VALIDATION_MAPPING
    row_start = 5

    if not regex:
        return None

    # Get the spreadsheet formula from the mapping or None if regex is not in the mapping
    # formula = REGEX_TO_SPREADSHEET_DATA_VALIDATION_MAPPING.get(regex, None)

    # # Replace placeholders with actual values
    # if formula:
    #     return formula.replace("{column_letter}", column_letter).replace(
    #         "{row_start}", str(row_start)
    #     )
    return None


def generate_json_file(data, output_file_path):
    """
    This function writes data to a JSON file.

    Parameters:
    data (dict): The data to write to the JSON file.
    output_file_path (str): The path to the JSON file.
    """
    directory_path = os.path.dirname(output_file_path)  # Get the directory path
    os.makedirs(
        directory_path, exist_ok=True
    )  # Create output directory if it does not exist
    file_name = os.path.basename(output_file_path)

    # Check if there's a conflicting directory with the same name as the file
    if os.path.isdir(output_file_path):
        print(
            f"Warning: A directory exists with the name '{output_file_path}'. Overwriting it."
        )
        shutil.rmtree(output_file_path)  # Remove the directory and its contents

    with open(output_file_path, "w") as f:
        f.write(json.dumps(data, indent=2, default=datetime_converter))

    print(f"'{file_name}' created!")


def generate_output_file_path(element, default_extension=DEFAULT_SCHEMA_EXTENSION):
    """
    Replace the default file extension in the given file path with a new extension
    incorporating namespace prefix.

    Parameters:
    - default_extension (str): Default file extension to replace. Default is '.xlsx'.
    - element (dict): A dictionary containing the following:
        data_df (pd.DataFrame): The input data frame.
        file_path (str): Original file path.
        output_file_name (str): Output file name to use.
        standard_name (str): Standard name to include in the new file name.
        technology_name (str): Schema name to include in the new file name.
        input_extension (str): Input file extension to search for. Default is '.json'.

    Returns:
    - str: Updated file path with the replaced extension.
    """
    data_df = element["data_df"]
    file_path = element["file_path"]
    output_file_name = element["output_file_name"]
    standard_name = element["standard_name"]
    technology_name = element["technology_name"]
    input_extension = element.get("input_extension", ".json")

    # Ensure file path ends with the input extension
    if not file_path.endswith(default_extension):
        raise ValueError(
            f"File path must end with {default_extension}, but got: {file_path}"
        )

    # Build the output file path incorporating the suffix i.e. the schema name
    output_directory = os.path.join(
        "dist", "checklists", input_extension.lstrip("."), standard_name
    )
    output_file_name = f"{output_file_name}{input_extension}"
    output_file_path = os.path.join(output_directory, output_file_name)
    return output_file_path


def get_required_columns(component_df, version_column_name):
    # Mandatory columns are columns with 'M' cells
    # Optional columns are columns with 'O' cells
    # Use the term name instead of term label as the key
    return component_df.loc[
        (component_df[version_column_name] == "M"), "term_name"
    ].tolist()


def get_checklists_from_xlsx_file():
    """
    Reads a spreadsheet file and extracts the checklists from the 'checklists' worksheet.
    The worksheet is expected to have 'key', 'name', 'standard' and 'technology' as column headers.

    Returns:
       - dict: A dictionary where keys are the 'key' column values, values are the
               'name' column values, standards as the 'standard' column values and
               technology as the 'technology' column values
    """
    # Declare the global variable
    global CHECKLISTS_DICT

    try:
        # Load the specified worksheet into a DataFrame
        checklists_df = pd.read_excel(SCHEMA_FILE_PATH, sheet_name="checklists").fillna(
            ""
        )

        # Ensure the required columns exist in the 'checklists' worksheet
        required_columns = {"key", "name", "description", "standard", "technology"}

        if not required_columns.issubset(checklists_df.columns):
            raise ValueError(
                f"'checklists' worksheet must contain {required_columns} columns."
            )

        # Create a dictionary where each 'key' maps to a dictionary of attributes
        checklists = {
            row["key"]: {
                "version_column_name": row["key"],
                "version_column_label": row["name"],
                "version_description": row["description"],
                "standard_name": row["standard"],
                "standard_label": re.search(r"\[(.*?)\]", row["name"]).group(1),
                "output_file_name": f"{row['technology']}_{row['standard']}",
                "technology_name": row["technology"],
                "technology_label": re.sub(r"\s*\[.*?\]", "", row["name"]).strip(),
            }
            for _, row in checklists_df.iterrows()
        }
        # Update the global dictionary
        CHECKLISTS_DICT.update(checklists)
    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{SCHEMA_FILE_PATH}' does not exist.")
    except Exception as e:
        raise RuntimeError(f"An error occurred while processing the file: {e}")


def merge_row(worksheet, row, last_column_letter, merge_format):
    """
    Function to merge cells in a row, unmerging any existing merged cells in that row if necessary.

    Args:
        worksheet: The worksheet object where the merge operation should be applied.
        row: The row number where merging is to occur.
        last_column_letter: The last column letter up to which the merge should occur.
        merge_format: The format to apply to the merged range.
    """
    # Iterate through all the existing merge ranges to check if any conflicts exist
    merged_range = None

    # Check if there are any merged ranges
    if worksheet.merged_cells.get("ranges", dict()):  # Check if there are merged ranges
        for merge_range in worksheet.merged_cells.ranges:
            # Parse the merge range in the format 'A1:B1'
            start_cell, end_cell = merge_range.split(":")
            start_row = int("".join(filter(str.isdigit, start_cell)))

            # If the row is the same as the one we're trying to merge
            if start_row == row:
                merged_range = (start_cell, end_cell)
                break

    try:
        # If the row is already merged, unmerge the conflicting range
        if merged_range:
            start_cell, end_cell = merged_range
            print(
                f"Row {row} is already merged between {start_cell} and {end_cell}. Undoing merge first."
            )
            worksheet.unmerge_range(f"{start_cell}:{end_cell}")

        # Proceed with merging the new range
        start_cell = f"A{row}"
        end_cell = f"{last_column_letter}{row}"

        if start_cell != end_cell:  # Ensure it's a valid range
            worksheet.merge_range(
                f"{start_cell}:{end_cell}",
                "FILL OUT INFORMATION BELOW THIS ROW",
                merge_format,
            )
        else:
            print(
                f"Skipping merge: {start_cell}:{end_cell}. Cannot merge a single cell because start and end cells are the same."
            )

    except Exception as e:
        print(f"Error: {e}")


def read_xlsx_data(return_dict=True):
    """
    Reads a spreadsheet file and returns a DataFrame and a dictionary of allowed values.

    Parameters:
    return_dict (bool): Flag to return a dictionary of allowed values. Default is True.

    Returns:
    tuple: A DataFrame containing the data sheet and a dictionary for allowed values.
    """
    # Declare the global variable
    global COMPONENTS
    global REGEX_TO_SPREADSHEET_DATA_VALIDATION_MAPPING

    try:
        # Load the spreadsheet file
        data_df = pd.read_excel(SCHEMA_FILE_PATH, sheet_name="data").fillna(
            ""
        )  # Replace NaN with empty strings
        allowed_values_df = pd.read_excel(
            SCHEMA_FILE_PATH, sheet_name="allowed_values", dtype=str
        )
        components_df = pd.read_excel(SCHEMA_FILE_PATH, sheet_name="components").fillna(
            ""
        )
        regex_to_formula_df = pd.read_excel(
            SCHEMA_FILE_PATH, sheet_name="regex_to_formula"
        ).fillna("")

        # 'data' worksheet logic
        # Strip whitespace from all string entries in the DataFrame
        data_df = data_df.apply(
            lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x)
        )

        # 'allowed_values' worksheet logic
        # Create a dictionary for allowed_values mapping
        allowed_values_dict = {
            column: allowed_values_df[column].dropna().tolist()  # Drop empty values
            for column in allowed_values_df.columns
        }

        # 'components' logic
        # Create a dictionary from the 'key' and 'name' columns
        components = dict(zip(components_df["key"], components_df["name"]))

        # Add the dictionary items directly to the global set as a frozenset
        COMPONENTS.update(frozenset(components.items()))

        # Regex logic
        # Create a dictionary from the 'regex' and 'formula' columns
        regex_to_formula = dict(
            zip(
                regex_to_formula_df["term_regex"],
                regex_to_formula_df["spreadsheet_formula"],
            )
        )
        REGEX_TO_SPREADSHEET_DATA_VALIDATION_MAPPING.update(
            frozenset(regex_to_formula.items())
        )

        # Validate if each regex in the term_regex column exists in the REGEX_TO_SPREADSHEET_DATA_VALIDATION_MAPPING
        validate_term_regex_values(data_df)

        # Return the DataFrame and the dictionary of allowed values based on the return_dict flag
        if return_dict:
            return data_df, allowed_values_dict
        else:
            return data_df, allowed_values_df
    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{SCHEMA_FILE_PATH}' does not exist.")
    except Exception as e:
        raise RuntimeError(f"An error occurred while processing the file: {e}")


def remove_dist_directory():
    """
    Remove the 'dist' directory and its contents.
    """
    dist_directory = "dist/checklists"
    if os.path.exists(dist_directory):
        shutil.rmtree(dist_directory)


def validate_allowed_values(row, allowed_values):
    if row["term_type"] == "enum" and not allowed_values:
        raise ValueError(
            f"Allowed values not found for '{row['term_name']}' in the 'allowed_values' sheet."
        )

    if row["term_type"] != "enum" and allowed_values:
        raise ValueError(
            f"Allowed values found for '{row['term_name']}' in the 'allowed_values' sheet, but term_type is not 'enum'."
        )


def validate_schema_file():
    file_name = os.path.basename(SCHEMA_FILE_PATH)

    if not file_name.startswith("base_") and not file_name.endswith(".xlsx"):
        raise ValueError(f"Unsupported file type: {SCHEMA_FILE_PATH}")
    return True


def validate_term_regex_values(data_df):
    global REGEX_TO_SPREADSHEET_DATA_VALIDATION_MAPPING

    error_messages = []

    # Filter rows where term_regex is not empty
    non_empty_regex_rows = data_df[data_df["term_regex"].astype(bool)]

    # Create a set of all regex values in the mapping for reverse lookup
    regex_in_map = {
        regex.encode("unicode_escape").decode("utf-8")
        for regex in REGEX_TO_SPREADSHEET_DATA_VALIDATION_MAPPING.keys()
    }

    for index, row in non_empty_regex_rows.iterrows():
        term_name = row["term_name"]
        term_regex_initial = row["term_regex"]

        # Normalise term_regex to match dictionary keys
        term_regex = term_regex_initial.encode("unicode_escape").decode("utf-8")

        # Check if the regex exists in the mapping
        if term_regex not in regex_in_map:
            # Raise ValueError with term_name and term_regex if it's not in the mapping
            error_messages.append(
                f"Regex '{term_regex_initial}' not found in 'regex_to_formula' worksheet for term '{term_name}'!"
            )

        # Check if any regex in the mapping is found in the dataframe
        if not any(map(lambda regex: regex in term_regex, regex_in_map)):
            error_messages.append(
                f"None of the regex patterns from the mapping found in the regex for term '{term_name}'!"
            )

    # If there are accumulated errors, raise a ValueError with all the messages
    if error_messages:
        raise ValueError("\n".join(error_messages))
    return True


def validate_version_column_value(data_df, non_empty, version_column_name):
    # Filter out only the non-empty rows
    non_empty_rows = data_df[non_empty]

    # Check if the valid values in the column are 'M' (mandatory) or 'O' (optional)
    is_version_column_value_valid = non_empty_rows[version_column_name].isin(["M", "O"])

    # Iterate through the rows where the validation fails
    invalid_rows = non_empty_rows[
        ~is_version_column_value_valid
    ]  # Get rows where validation fails

    if not invalid_rows.empty:
        for index, row in invalid_rows.iterrows():
            # You can specify what exactly you want to report (invalid value, row number, and column name)
            invalid_value = row[version_column_name]
            raise ValueError(
                f"Invalid value '{invalid_value}' found in row {index} of column '{version_column_name}'. Only 'M' or 'O' are allowed."
            )
