'''
Script Name: update_and_convert_schema.py
Description:
    This script updates the information in the 'data' worksheet of a base schema
    Excel file and propagates changes to other worksheets. The updated Excel file
    is then converted into JSON format for further use.

Usage:
    python update_and_convert_schema.py --file_path <path_to_file> --output <output_path>
'''

import json
import os
import pandas as pd
import tempfile
import utils.helpers as helpers

from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.utils import get_column_letter

def autofit_all_sheets(writer):
    for sheet in writer.sheets.values():
        sheet.autofit()

def generate_base_schema_xlsx(file_path):
    data_df, allowed_values_df = helpers.read_excel_data(file_path, return_dict=False)
    
    # Create a temporary file to write the output
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_output_path = temp_file.name
        
    # Create the Excel writer object
    with pd.ExcelWriter(temp_output_path, engine='xlsxwriter') as writer:
        # Write the 'data' worksheet and 'allowed_values' worksheet
        data_df.to_excel(writer, sheet_name='data', index=False)
        allowed_values_df.to_excel(writer, sheet_name='allowed_values', index=False)
        
        # Group data by namespace prefix and write each namespace prefix to its own protected sheet
        for namespace_prefix, group in data_df.groupby('namespace_prefix'):
            namespace_prefix_df = group.copy()
            sheet_name = namespace_prefix[:31]  # Ensure sheet name length is valid for Excel
            namespace_prefix_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Autofit columns for all sheets
        autofit_all_sheets(writer)
    # Protect namespace prefix sheets
    wb = load_workbook(temp_output_path)
    for sheet in wb.worksheets:
        if sheet.title not in ['data', 'allowed_values']:
            for col in range(1, sheet.max_column + 1):
                col_letter = get_column_letter(col)
                sheet.column_dimensions[col_letter].hidden = False
            sheet.protection = SheetProtection(sheet=True)
    
    # Save the workbook with protections applied
    wb.save(temp_output_path)
    
    # Overwrite the original file with the generated file
    os.replace(temp_output_path, file_path)
    print(f'Excel file updated: {file_path}')
    
def generate_base_schema_json(file_path):
    data_df, allowed_values_dict = helpers.read_excel_data(file_path)
    
    json_data = helpers.get_base_schema_json(data_df, allowed_values_dict)

    # Write JSON to a temporary file
    with tempfile.NamedTemporaryFile(suffix=".json", delete=False, mode="w") as temp_file:
        temp_output_path = temp_file.name
        json.dump(json_data, temp_file, indent=4)
    
    # Overwrite the original file with the generated JSON file
    file_name = os.path.basename(file_path).replace('.xlsx', '.json')
    output_file = f'{helpers.SCHEMA_BASE_DIR_PATH_JSON}/{file_name}'
    os.replace(temp_output_path, output_file)
    print(f'JSON file generated/updated: {output_file}')
    
if __name__ == '__main__':    
    for file_path in helpers.SCHEMA_FILE_PATHS:
        if file_path.endswith('.xlsx'):
            generate_base_schema_xlsx(file_path)
            generate_base_schema_json(file_path)
            print('\n_______\n')
        else:
            print(f'Unsupported file type: {file_path}')